import openpyxl


#same type of data
# file="C:\\Users\\HP\\Downloads\\sample3.xlsx"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook["Sheet1"]

# for r in range(1,11):
#     for c in range(1,5):
#         sheet.cell(r,c).value="Hemanth"

# workbook.save(file)


#Multiple type of data
file="C:\\Users\\HP\\Downloads\\sample4.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

sheet.cell(1,1).value=11
sheet.cell(1,2).value="Hemanth"

sheet.cell(2,1).value=12
sheet.cell(2,2).value="Vishnu"

sheet.cell(3,1).value=13
sheet.cell(3,2).value="Sankar"

sheet.cell(4,1).value=14
sheet.cell(4,2).value="Nirmala"

workbook.save(file)