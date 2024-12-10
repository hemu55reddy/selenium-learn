import openpyxl


file="C:\\Users\\HP\\Downloads\\sample1.xlsx"
workbook=openpyxl.load_workbook(file) #store the file in workbook
sheet=workbook["Sheet1"] #store the workbook in sheet

rows=sheet.max_row #read the number of/ max rows
print(rows)
columns=sheet.max_column #read the number of/ max columns
print(columns)


#read all the c and r from the sheet
for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value, end='     ')
    print()


#https://filesamples.com/formats/xlsx#google_vignette  (example file for download )